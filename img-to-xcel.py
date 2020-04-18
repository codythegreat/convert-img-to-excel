#!/usr/bin/env python3

from PIL import Image
import numpy as np
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# open the image and create a numpy
# array of each pixel's RGB value
img = Image.open(sys.argv[1])
arr = np.array(img)

# create an empty workbook
wb = Workbook()
ws = wb.active


# iterate over each pixel's RGB value
# convert each value to hexidecimal
# use hex to create bachground color
col = 1
row = 1
for x in arr:
    for value in x:
        red = f"{value[0]:x}".upper().rjust(2, "0") + "0000" 
        green = "00" + f"{value[1]:x}".upper().rjust(2, "0") + "00"
        blue = "0000" + f"{value[2]:x}".upper().rjust(2, "0")

        ws.cell(row=row, column=col).fill = PatternFill(
            start_color=red,
            end_color=red,
            fill_type="solid"
        )
        ws.cell(row=row+1, column=col).fill = PatternFill(
            start_color=green,
            end_color=green,
            fill_type="solid"
        )
        ws.cell(row=row+2, column=col).fill = PatternFill(
            start_color=blue,
            end_color=blue,
            fill_type="solid"
        )

        col += 1
    
    # itterate to next three rows and reset to first column
    row += 3
    col = 1

# save the excel file
wb.save(filename='out.xlsx')